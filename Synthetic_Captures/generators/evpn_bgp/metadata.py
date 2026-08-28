"""Metadata writer — generates dataset_metadata.xlsx."""

import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .config import TopologyConfig


# Styling constants
_HEADER_FONT = Font(bold=True, size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
_ALT_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _auto_column_widths(ws):
    """Adjust column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)


def _style_header_row(ws, row_num: int, num_cols: int):
    """Apply header styling to a row."""
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = _HEADER_FONT_WHITE
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _THIN_BORDER


def _style_data_rows(ws, start_row: int, end_row: int, num_cols: int):
    """Apply alternating row colors and borders."""
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if (row_idx - start_row) % 2 == 1:
                cell.fill = _ALT_FILL


def _fn(fault_type: str, variant: str | None) -> str:
    """Build filename matching _filename_for_scenario with copy_idx=1."""
    parts = fault_type.replace("-", "_")
    if variant:
        parts += f"_{variant.replace('-', '_')}"
    return f"{parts}.pcap"


class MetadataWriter:
    """Generates Excel metadata file documenting the synthetic dataset."""

    def __init__(self, config: TopologyConfig, output_dir: Path):
        self.config = config
        self.output_dir = Path(output_dir)
        self.generated_files = []  # List of dicts with file info

    def record_file(self, filename: str, section: int, fault_type: str = None,
                    variant: str = None, affected_device: str = None,
                    fault_time: str = None, recovery: bool = None,
                    recovery_time: str = None, description: str = "",
                    ground_truth: str = None):
        """Record metadata for a generated file."""
        self.generated_files.append({
            'filename': filename,
            'section': section,
            'fault_type': fault_type,
            'variant': variant,
            'affected_device': affected_device,
            'fault_time': fault_time,
            'recovery': recovery,
            'recovery_time': recovery_time,
            'description': description,
            'ground_truth': ground_truth,
        })

    def write(self, output_path: str | Path = None):
        """Write the metadata Excel file."""
        path = Path(output_path) if output_path else self.output_dir / "dataset_metadata.xlsx"
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        self._write_global_config(wb)
        self._write_section1_labels(wb)
        self._write_section2_labels(wb)
        self._write_section3_labels(wb)
        self._write_section4_labels(wb)

        # Remove default empty sheet if created
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(str(path))
        return path

    def _write_global_config(self, wb: Workbook):
        """Write Global Config sheet."""
        ws = wb.create_sheet("Global Config")

        rows = [
            ("Parameter", "Value"),
            ("AS Number", self.config.as_number),
            ("Hold Timer", f"{self.config.timing.hold_timer}s"),
            ("Keepalive Timer", f"{self.config.timing.keepalive_timer}s"),
            ("PEs", len(self.config.pe_nodes)),
            ("RRs", len(self.config.route_reflectors)),
            ("EVPN VNI", self.config.evpn.vni),
            ("Route Target", self.config.evpn.route_target),
            ("Transport", "IPv6"),
            ("Encapsulation", "SRv6"),
            ("Link Type", "CookedLinuxV2 (276)"),
            ("Capture Vantage", self.config.capture_vantage),
        ]

        for row_idx, (param, value) in enumerate(rows, start=1):
            ws.cell(row=row_idx, column=1, value=param)
            ws.cell(row=row_idx, column=2, value=value)

        _style_header_row(ws, 1, 2)
        _style_data_rows(ws, 2, len(rows), 2)
        for row_idx in range(2, len(rows) + 1):
            ws.cell(row=row_idx, column=1).font = _HEADER_FONT

        _auto_column_widths(ws)

    def _write_section1_labels(self, wb: Workbook):
        """Write Section 1 Labels sheet — normal baseline traffic."""
        ws = wb.create_sheet("Section 1 Labels")

        headers = ["Filename", "Traffic Type", "Load Profile", "Description"]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        section1_files = [f for f in self.generated_files if f['section'] == 1]

        for row_idx, entry in enumerate(section1_files, start=2):
            ws.cell(row=row_idx, column=1, value=entry['filename'])
            ws.cell(row=row_idx, column=2, value=entry['fault_type'] or "normal")
            ws.cell(row=row_idx, column=3, value=entry['variant'] or "default")
            ws.cell(row=row_idx, column=4, value=entry['description'])

        num_cols = len(headers)
        _style_header_row(ws, 1, num_cols)
        _style_data_rows(ws, 2, len(section1_files) + 1, num_cols)
        _auto_column_widths(ws)

    def _write_section2_labels(self, wb: Workbook):
        """Write Section 2 Labels sheet — single-fault labelled scenarios."""
        ws = wb.create_sheet("Section 2 Labels")

        headers = [
            "Filename", "Fault Type", "Variant", "Affected Device",
            "Fault Inject Time (s)", "Recovery", "Recovery Time (s)", "Description"
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        section2_files = [f for f in self.generated_files if f['section'] == 2]

        for row_idx, entry in enumerate(section2_files, start=2):
            ws.cell(row=row_idx, column=1, value=entry['filename'])
            ws.cell(row=row_idx, column=2, value=entry['fault_type'] or "")
            ws.cell(row=row_idx, column=3, value=entry['variant'] or "")
            ws.cell(row=row_idx, column=4, value=entry['affected_device'] or "")
            ws.cell(row=row_idx, column=5, value=entry['fault_time'] if entry['fault_time'] else "")
            ws.cell(row=row_idx, column=6, value="Yes" if entry['recovery'] else "No")
            ws.cell(row=row_idx, column=7, value=entry['recovery_time'] if entry['recovery_time'] else "N/A")
            ws.cell(row=row_idx, column=8, value=entry['description'])

        num_cols = len(headers)
        _style_header_row(ws, 1, num_cols)
        _style_data_rows(ws, 2, len(section2_files) + 1, num_cols)
        _auto_column_widths(ws)

    def _write_section3_labels(self, wb: Workbook):
        """Write Section 3 Labels sheet — mixed/eval inference scenarios."""
        ws = wb.create_sheet("Section 3 Labels")

        headers = [
            "Filename", "Scenario Type", "Faults Present",
            "Fault Inject Time(s)", "Recovery", "Ground Truth Label", "Description"
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        section3_files = [f for f in self.generated_files if f['section'] == 3]

        for row_idx, entry in enumerate(section3_files, start=2):
            ws.cell(row=row_idx, column=1, value=entry['filename'])
            ws.cell(row=row_idx, column=2, value=entry['variant'] or "")
            ws.cell(row=row_idx, column=3, value=entry['fault_type'] or "none")
            ws.cell(row=row_idx, column=4, value=entry['fault_time'] if entry['fault_time'] else "N/A")
            ws.cell(row=row_idx, column=5, value="Yes" if entry['recovery'] else "No")
            ws.cell(row=row_idx, column=6, value=entry['ground_truth'] or "")
            ws.cell(row=row_idx, column=7, value=entry['description'])

        num_cols = len(headers)
        _style_header_row(ws, 1, num_cols)
        _style_data_rows(ws, 2, len(section3_files) + 1, num_cols)
        _auto_column_widths(ws)

    def _write_section4_labels(self, wb: Workbook):
        """Write Section 4 Labels sheet — additional temporal/cascade scenarios."""
        ws = wb.create_sheet("Section 4 Labels")

        headers = [
            "Filename", "Scenario Type", "Faults Present",
            "Fault Inject Time(s)", "Recovery", "Ground Truth Label", "Description"
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        section4_files = [f for f in self.generated_files if f['section'] == 4]

        for row_idx, entry in enumerate(section4_files, start=2):
            ws.cell(row=row_idx, column=1, value=entry['filename'])
            ws.cell(row=row_idx, column=2, value=entry['variant'] or "")
            ws.cell(row=row_idx, column=3, value=entry['fault_type'] or "none")
            ws.cell(row=row_idx, column=4, value=entry['fault_time'] if entry['fault_time'] else "N/A")
            ws.cell(row=row_idx, column=5, value="Yes" if entry['recovery'] else "No")
            ws.cell(row=row_idx, column=6, value=entry['ground_truth'] or "")
            ws.cell(row=row_idx, column=7, value=entry['description'])

        num_cols = len(headers)
        _style_header_row(ws, 1, num_cols)
        _style_data_rows(ws, 2, len(section4_files) + 1, num_cols)
        _auto_column_widths(ws)


def _warmup_for(cls_path: str) -> str:
    """Return a human-readable fault-inject window for a scenario class.

    WARMUP_SECONDS is now a (min, max) tuple — returns e.g. '120–480s (variable)'.
    Falls back to '120–480s (variable)' if the class cannot be loaded.
    """
    try:
        import importlib
        parts = cls_path.rsplit(".", 1)
        mod = importlib.import_module(parts[0])
        cls = getattr(mod, parts[1])
        ws = getattr(cls, "WARMUP_SECONDS", (120, 480))
        if isinstance(ws, tuple):
            return f"{ws[0]}–{ws[1]}s (variable)"
        return f"~{ws}s"
    except Exception:
        return "120–480s (variable)"


def generate_default_metadata(config: TopologyConfig, output_dir: Path) -> MetadataWriter:
    """Generate metadata for the standard full dataset (all 165 scenarios).

    Returns a pre-populated MetadataWriter covering:
      Section 1 —  9 normal baseline traffic files
      Section 2 — 62 single-fault labelled scenarios
      Section 3 — 73 mixed/eval inference scenarios
      Section 4 — 21 additional temporal/cascade scenarios
    """
    writer = MetadataWriter(config, output_dir)

    # Import SCENARIO_REGISTRY so we can look up the actual class for each variant
    from .cli import SCENARIO_REGISTRY

    def _ft(section: int, fault_type: str, variant: str | None) -> int:
        """Fault injection time (s) read directly from the scenario class.

        Falls back to 300 and prints a warning if no matching entry exists
        in cli.py's SCENARIO_REGISTRY.
        """
        sec = SCENARIO_REGISTRY.get(section, {})
        ft_map = sec.get(fault_type, {})
        cls_path = ft_map.get(variant)
        if cls_path:
            return _warmup_for(cls_path)
        print(
            f"WARNING: metadata.py._ft() found no SCENARIO_REGISTRY entry for "
            f"(section={section}, fault_type={fault_type!r}, variant={variant!r}) "
            f"-- falling back to fault_time=300. This scenario's metadata row "
            f"will carry a placeholder fault_time; check whether this "
            f"(section, fault_type, variant) still exists in cli.py's "
            f"SCENARIO_REGISTRY or whether this record_file() call is stale "
            f"and needs updating/removing.",
            file=sys.stderr,
        )
        return 300  # safe default

    # =========================================================================
    # SECTION 1 — Normal baseline traffic (9 files)
    # =========================================================================

    # writer.record_file(
        # filename=_fn("quiet", None), section=1,
        # fault_type="quiet", variant="default",
        # description="Quiet baseline EVPN fabric: all 5 PEs + 2 RRs, low-load keepalive + IMET/MAC-IP routes, ~116K frames",
    # )
    # writer.record_file(
        # filename=_fn("quiet", "pe1-pe3"), section=1,
        # fault_type="quiet", variant="pe1-pe3",
        # description="Quiet baseline traffic captured from PE1-PE3 vantage perspective",
    # )
    # writer.record_file(
        # filename=_fn("quiet", "pe4-pe5"), section=1,
        # fault_type="quiet", variant="pe4-pe5",
        # description="Quiet baseline traffic captured from PE4-PE5 vantage perspective",
    # )
    # writer.record_file(
        # filename=_fn("moderate", None), section=1,
        # fault_type="moderate", variant="default",
        # description="Moderate-load EVPN fabric: standard keepalive interval, regular UPDATE churn, ~121K frames",
    # )
    # writer.record_file(
        # filename=_fn("moderate", "pe2-pe4"), section=1,
        # fault_type="moderate", variant="pe2-pe4",
        # description="Moderate traffic captured from PE2-PE4 vantage perspective",
    # )
    # writer.record_file(
        # filename=_fn("moderate", "pe1-pe5"), section=1,
        # fault_type="moderate", variant="pe1-pe5",
        # description="Moderate traffic captured from PE1-PE5 vantage perspective",
    # )
    # writer.record_file(
        # filename=_fn("busy", None), section=1,
        # fault_type="busy", variant="default",
        # description="High-load EVPN fabric: dense keepalive + frequent EVPN UPDATE churn, ~114K frames",
    # )
    # writer.record_file(
        # filename=_fn("busy", "pe2-pe3"), section=1,
        # fault_type="busy", variant="pe2-pe3",
        # description="Busy traffic captured from PE2-PE3 vantage perspective",
    # )
    # writer.record_file(
        # filename=_fn("busy", "pe1-pe4"), section=1,
        # fault_type="busy", variant="pe1-pe4",
        # description="Busy traffic captured from PE1-PE4 vantage perspective",
    # )

    # =========================================================================
    # SECTION 2 — Single-fault labelled scenarios (62 files)
    # =========================================================================

    # --- Link-Down: Fast Recovery (5 files) ---
    # TCP RST tears session, reconnects in 20-30s — no BGP NOTIFICATION emitted
    # for pe in range(1, 6):
        # ft = _ft(2, "link-down", f"fast-recovery-pe{pe}")
        # writer.record_file(
            # filename=_fn("link-down", f"fast-recovery-pe{pe}"), section=2,
            # fault_type="link-down", variant=f"fast-recovery-pe{pe}",
            # affected_device=f"PE{pe}",
            # fault_time=ft, recovery=True, recovery_time="~25s after fault",
            # description=(
                # f"PE{pe}-RR link drops abruptly; TCP RST clears the BGP session immediately. "
                # f"Session reconnects within 20-30s. No BGP NOTIFICATION — TCP teardown only."
            # ),
        # )

    # --- Link-Down: Slow Recovery (5 files) ---
    # Hold timer expires before reconnect — BGP NOTIFICATION code 4 (Hold Timer Expired)
    # for pe in range(1, 6):
        # ft = _ft(2, "link-down", f"slow-recovery-pe{pe}")
        # writer.record_file(
            # filename=_fn("link-down", f"slow-recovery-pe{pe}"), section=2,
            # fault_type="link-down", variant=f"slow-recovery-pe{pe}",
            # affected_device=f"PE{pe}",
            # fault_time=ft, recovery=True, recovery_time="~90s after fault",
            # description=(
                # f"PE{pe}-RR link drops silently; hold timer expires after ~90s triggering a "
                # f"BGP NOTIFICATION (Hold Timer Expired). Session eventually re-establishes."
            # ),
        # )

    # --- Link-Down: No Recovery (5 files) ---
    # Link lost permanently; hold timer fires, NOTIFICATION sent, no reconnect
    # for pe in range(1, 6):
        # ft = _ft(2, "link-down", f"no-recovery-pe{pe}")
        # writer.record_file(
            # filename=_fn("link-down", f"no-recovery-pe{pe}"), section=2,
            # fault_type="link-down", variant=f"no-recovery-pe{pe}",
            # affected_device=f"PE{pe}",
            # fault_time=ft, recovery=False, recovery_time=None,
            # description=(
                # f"PE{pe}-RR link drops permanently; hold timer expires → "
                # f"BGP NOTIFICATION (Hold Timer Expired). Session never re-establishes. "
                # f"Capture ends with PE{pe} unreachable."
            # ),
        # )

    # --- Link-Down: Hold Timer Expiry (5 files) ---
    # Keepalive loss causes hold timer to tick down and expire
    # for pe in range(1, 6):
        # ft = _ft(2, "link-down", f"hold-timer-pe{pe}")
        # writer.record_file(
            # filename=_fn("link-down", f"hold-timer-pe{pe}"), section=2,
            # fault_type="link-down", variant=f"hold-timer-pe{pe}",
            # affected_device=f"PE{pe}",
            # fault_time=ft, recovery=False, recovery_time=None,
            # description=(
                # f"PE{pe} keepalives stop being received; hold timer counts to zero → "
                # f"BGP NOTIFICATION (Hold Timer Expired). Link remains down."
            # ),
        # )

    # --- Link-Down: Simultaneous (1 file) ---
    # ft = _ft(2, "link-down", "simultaneous")
    # writer.record_file(
        # filename=_fn("link-down", "simultaneous"), section=2,
        # fault_type="link-down", variant="simultaneous",
        # affected_device="PE1,PE2",
        # fault_time=ft, recovery=True, recovery_time="~35s after fault",
        # description=(
            # "PE1 and PE2 links drop simultaneously via TCP RST; both sessions reconnect. "
            # "Tests multi-session concurrent failure detection."
        # ),
    # )

    # --- RR-Down: Clean Restart (2 files) ---
    # for rr in (1, 2):
        # ft = _ft(2, "rr-down", f"clean-restart-rr{rr}")
        # writer.record_file(
            # filename=_fn("rr-down", f"clean-restart-rr{rr}"), section=2,
            # fault_type="rr-down", variant=f"clean-restart-rr{rr}",
            # affected_device=f"RR{rr}",
            # fault_time=ft, recovery=True, recovery_time="~28s after fault",
            # description=(
                # f"RR{rr} shuts down cleanly; all 5 PE sessions drop via TCP RST. "
                # f"RR{rr} restarts and sessions re-establish within 25-30s. "
                # f"High OPEN count from mass re-establishment."
            # ),
        # )

    # --- RR-Down: Slow Restart (2 files) ---
    # for rr in (1, 2):
        # ft = _ft(2, "rr-down", f"slow-restart-rr{rr}")
        # writer.record_file(
            # filename=_fn("rr-down", f"slow-restart-rr{rr}"), section=2,
            # fault_type="rr-down", variant=f"slow-restart-rr{rr}",
            # affected_device=f"RR{rr}",
            # fault_time=ft, recovery=True, recovery_time="~90s after fault",
            # description=(
                # f"RR{rr} goes down; hold timers expire on all PEs → multiple BGP NOTIFICATIONs "
                # f"(Hold Timer Expired). RR{rr} eventually restarts and sessions recover."
            # ),
        # )

    # --- RR-Down: No Recovery (2 files) ---
    # for rr in (1, 2):
        # ft = _ft(2, "rr-down", f"no-recovery-rr{rr}")
        # writer.record_file(
            # filename=_fn("rr-down", f"no-recovery-rr{rr}"), section=2,
            # fault_type="rr-down", variant=f"no-recovery-rr{rr}",
            # affected_device=f"RR{rr}",
            # fault_time=ft, recovery=False, recovery_time=None,
            # description=(
                # f"RR{rr} goes down permanently; all PE sessions are lost. "
                # f"Hold timers expire for each PE → BGP NOTIFICATIONs. No recovery."
            # ),
        # )

    # --- RR-Down: Both Simultaneous (1 file) ---
    # ft = _ft(2, "rr-down", "both-simultaneous")
    # writer.record_file(
        # filename=_fn("rr-down", "both-simultaneous"), section=2,
        # fault_type="rr-down", variant="both-simultaneous",
        # affected_device="RR1,RR2",
        # fault_time=ft, recovery=False, recovery_time=None,
        # description=(
            # "Both RR1 and RR2 go down simultaneously; all PE-RR sessions are lost. "
            # "Complete route-reflector outage scenario."
        # ),
    # )

    # --- ESDF-Toggle: Single (5 files) ---
    for pe in (1, 2):
        ft = _ft(2, "esdf-toggle", f"single-pe{pe}")
        writer.record_file(
            filename=_fn("esdf-toggle", f"single-pe{pe}"), section=2,
            fault_type="esdf-toggle", variant=f"single-pe{pe}",
            affected_device=f"PE{pe}",
            fault_time=ft, recovery=True, recovery_time="~15s after fault",
            description=(
                f"Single ES/DF election toggle on PE{pe}: EAD/ES routes withdrawn then "
                f"re-advertised within 10-20s. Clean recovery, no BGP session drop."
            ),
        )

    # --- ESDF-Toggle: Rapid (5 files) ---
    for pe in (1, 2):
        ft = _ft(2, "esdf-toggle", f"rapid-pe{pe}")
        writer.record_file(
            filename=_fn("esdf-toggle", f"rapid-pe{pe}"), section=2,
            fault_type="esdf-toggle", variant=f"rapid-pe{pe}",
            affected_device=f"PE{pe}",
            fault_time=ft, recovery=True, recovery_time="~45s after fault",
            description=(
                f"Rapid ES/DF flapping on PE{pe}: multiple EAD/ES withdraw + re-advertise "
                f"cycles in quick succession. High UPDATE count from continuous churn."
            ),
        )

    # --- ESDF-Toggle: No Recovery (5 files) ---
    for pe in (1, 2):
        ft = _ft(2, "esdf-toggle", f"no-recovery-pe{pe}")
        writer.record_file(
            filename=_fn("esdf-toggle", f"no-recovery-pe{pe}"), section=2,
            fault_type="esdf-toggle", variant=f"no-recovery-pe{pe}",
            affected_device=f"PE{pe}",
            fault_time=ft, recovery=False, recovery_time=None,
            description=(
                f"ES/DF toggle on PE{pe}: EAD/ES routes withdrawn and never re-advertised. "
                f"Ethernet Segment silently disappears from fabric. No BGP session drop."
            ),
        )

    # --- ESDF-Toggle: Slow (1 file) ---
    ft = _ft(2, "esdf-toggle", "slow")
    writer.record_file(
        filename=_fn("esdf-toggle", "slow"), section=2,
        fault_type="esdf-toggle", variant="slow",
        affected_device="PE1",
        fault_time=ft, recovery=True, recovery_time="~120s after fault",
        description=(
            "Slow ES/DF toggle: extended dwell time between EAD/ES withdraw and re-advertise. "
            "Tests detection of low-frequency ES churn."
        ),
    )

    # --- ESDF-Toggle: Type-1 per-EVI EAD trigger (2 files) ---
    for pe in (1, 2):
        ft = _ft(2, "esdf-toggle", f"type1-evi-pe{pe}")
        writer.record_file(
            filename=_fn("esdf-toggle", f"type1-evi-pe{pe}"), section=2,
            fault_type="esdf-toggle", variant=f"type1-evi-pe{pe}",
            affected_device=f"PE{pe}",
            fault_time=ft, recovery=True, recovery_time="~15s after fault",
            description=(
                f"ES/DF election toggle on PE{pe} triggered by a Type-1 per-EVI EAD route "
                f"withdrawal (RFC 8584's second DF-election trigger type), then "
                f"re-advertised within 10-20s. Clean recovery, no BGP session drop."
            ),
        )

    # --- ESDF-Toggle: Local AC state trigger (2 files) ---
    for pe in (1, 2):
        ft = _ft(2, "esdf-toggle", f"ac-state-pe{pe}")
        writer.record_file(
            filename=_fn("esdf-toggle", f"ac-state-pe{pe}"), section=2,
            fault_type="esdf-toggle", variant=f"ac-state-pe{pe}",
            affected_device=f"PE{pe}",
            fault_time=ft, recovery=True, recovery_time="~15s after fault",
            description=(
                f"ES/DF election toggle on PE{pe} triggered by local AC state "
                f"(RFC 8584's first DF-election trigger type): the DF Election "
                f"Extended Community's AC-DF bit is cleared on a Type-4 re-advertisement, "
                f"then set again 10-20s later. No route withdrawal at any point."
            ),
        )

    # --- RT-Misconfig: Wrong RT export (5 files) ---
    # for pe in range(1, 6):
        # ft = _ft(2, "rt-misconfig", f"pe{pe}")
        # writer.record_file(
            # filename=_fn("rt-misconfig", f"pe{pe}"), section=2,
            # fault_type="rt-misconfig", variant=f"pe{pe}",
            # affected_device=f"PE{pe}",
            # fault_time=ft, recovery=False, recovery_time=None,
            # description=(
                # f"PE{pe} advertises all EVPN routes with a wrong Route Target value. "
                # f"Routes are silently dropped by all peers — no BGP session impact, "
                # f"no NOTIFICATION. Pure control-plane misconfiguration."
            # ),
        # )

    # --- RT-Misconfig: Import mismatch (5 files) ---
#     for pe in range(1, 6):
#         ft = _ft(2, "rt-misconfig", f"import-pe{pe}")
#         writer.record_file(
#             filename=_fn("rt-misconfig", f"import-pe{pe}"), section=2,
#             fault_type="rt-misconfig", variant=f"import-pe{pe}",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=False, recovery_time=None,
#             description=(
#                 f"PE{pe} has an import RT mismatch: peers' routes are received but silently "
#                 f"discarded because the import policy does not match the advertised RT."
#             ),
#         )

    # --- RT-Misconfig: Export mismatch (5 files) ---
#     for pe in range(1, 6):
#         ft = _ft(2, "rt-misconfig", f"export-pe{pe}")
#         writer.record_file(
#             filename=_fn("rt-misconfig", f"export-pe{pe}"), section=2,
#             fault_type="rt-misconfig", variant=f"export-pe{pe}",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=False, recovery_time=None,
#             description=(
#                 f"PE{pe} exports EVPN routes with an incorrect RT in MP_REACH_NLRI. "
#                 f"All peers silently drop the routes; no session disruption."
#             ),
#         )

    # recovery is covered by the real pilot_containerlab testbed, same as
    # the plain export variant above. Commented out to match cli.py's
    # SCENARIO_REGISTRY, not deleted.
    # --- RT-Misconfig: With Recovery (3 files — pe1, pe2, pe4 only) ---
    # for pe in (1, 2, 4):
        # ft = _ft(2, "rt-misconfig", f"recovery-pe{pe}")
        # writer.record_file(
            # filename=_fn("rt-misconfig", f"recovery-pe{pe}"), section=2,
            # fault_type="rt-misconfig", variant=f"recovery-pe{pe}",
            # affected_device=f"PE{pe}",
            # fault_time=ft, recovery=True, recovery_time="~60s after fault",
            # description=(
                # f"PE{pe} starts with wrong RT export; misconfiguration corrected mid-capture. "
                # f"Routes are re-advertised with the correct RT and peers accept them."
            # ),
        # )

    # --- ESDF-Toggle: Full Failure (2 files) ---
    ft = _ft(2, "esdf-toggle", "full-failure-recovery")
    writer.record_file(
        filename=_fn("esdf-toggle", "full-failure-recovery"), section=2,
        fault_type="esdf-toggle", variant="full-failure-recovery",
        affected_device="PE1,PE2",
        fault_time=ft, recovery=True, recovery_time="~15s after fault",
        description=(
            "Both multihomed PEs (PE1, PE2) lose DF role together: ES routes "
            "withdrawn on both, then re-advertised. Clean recovery."
        ),
    )
    ft = _ft(2, "esdf-toggle", "full-failure-no-recovery")
    writer.record_file(
        filename=_fn("esdf-toggle", "full-failure-no-recovery"), section=2,
        fault_type="esdf-toggle", variant="full-failure-no-recovery",
        affected_device="PE1,PE2",
        fault_time=ft, recovery=False, recovery_time=None,
        description=(
            "Both multihomed PEs (PE1, PE2) lose DF role together: ES routes "
            "withdrawn on both and never re-advertised. No recovery."
        ),
    )

    # --- RT-Misconfig: ES-Import (4 files) ---
    for pe in (1, 2):
        ft = _ft(2, "rt-misconfig", f"es-import-pe{pe}")
        writer.record_file(
            filename=_fn("rt-misconfig", f"es-import-pe{pe}"), section=2,
            fault_type="rt-misconfig", variant=f"es-import-pe{pe}",
            affected_device=f"PE{pe}",
            fault_time=ft, recovery=False, recovery_time=None,
            description=(
                f"PE{pe} advertises its Type-4 ES route with a wrong Route Target; "
                f"Type-2 MAC/IP routes from the same PE are unaffected. Breaks "
                f"ES-Import RT matching / DF election discovery between the ESI peers."
            ),
        )
    for pe in (1, 2):
        ft = _ft(2, "rt-misconfig", f"es-import-recovery-pe{pe}")
        writer.record_file(
            filename=_fn("rt-misconfig", f"es-import-recovery-pe{pe}"), section=2,
            fault_type="rt-misconfig", variant=f"es-import-recovery-pe{pe}",
            affected_device=f"PE{pe}",
            fault_time=ft, recovery=True, recovery_time="~60s after fault",
            description=(
                f"PE{pe} starts with a wrong RT on its Type-4 ES route; misconfiguration "
                f"corrected mid-capture and the ES route is re-advertised with the correct RT."
            ),
        )

    # --- MAC Mobility: Rapid Flap (4 files) ---
    for pair in ("pe1-pe2", "pe2-pe1"):
        ft = _ft(2, "mac-mobility", f"rapid-{pair}")
        writer.record_file(
            filename=_fn("mac-mobility", f"rapid-{pair}"), section=2,
            fault_type="mac-mobility", variant=f"rapid-{pair}",
            affected_device=pair.upper().replace("-", ","),
            fault_time=ft, recovery=False, recovery_time=None,
            description=(
                f"Single MAC Mobility rapid-flap event ({pair}): WITHDRAW from the "
                f"old-owner PE, then ADVERTISE from the new-owner PE with an "
                f"incremented RFC 7432 SS15 sequence number, 2.0s apart -- inside "
                f"the confirmed real testbed clean-move delta range of 1.74-6.70s."
            ),
        )

    # --- MAC Mobility: Repeated Flap (4 files) ---
    for pair in ("pe1-pe2", "pe2-pe1"):
        ft = _ft(2, "mac-mobility", f"repeated-{pair}")
        writer.record_file(
            filename=_fn("mac-mobility", f"repeated-{pair}"), section=2,
            fault_type="mac-mobility", variant=f"repeated-{pair}",
            affected_device=pair.upper().replace("-", ","),
            fault_time=ft, recovery=False, recovery_time=None,
            description=(
                f"MAC Mobility flap storm ({pair}): the same MAC moves back and "
                f"forth 3-6 times within one capture, each flap using the same "
                f"WITHDRAW-then-ADVERTISE (2.0s gap) ordering as the single-flap "
                f"variant. Sequence number increments monotonically across the "
                f"whole capture, never resetting per flap."
            ),
        )

    # =========================================================================
    # SECTION 3 — Mixed / eval inference scenarios (73 files)
    # =========================================================================

    # --- Overlapping faults (3 files) ---
    # ft = _ft(3, "overlapping", "ld-ld-pe2-pe3")
    # writer.record_file(
        # filename=_fn("overlapping", "ld-ld-pe2-pe3"), section=3,
        # fault_type="link-down,link-down", variant="overlapping",
        # affected_device="PE2,PE3",
        # fault_time=ft, recovery=False,
        # ground_truth="ANOMALY→COMPOUND",
        # description="PE2 and PE3 links drop in overlapping windows; both sessions fail simultaneously.",
    # )
    # ft = _ft(3, "overlapping", "ld-rr-pe1-rr2")
    # writer.record_file(
        # filename=_fn("overlapping", "ld-rr-pe1-rr2"), section=3,
        # fault_type="link-down,rr-down", variant="overlapping",
        # affected_device="PE1,RR2",
        # fault_time=ft, recovery=True, recovery_time="~45s after fault",
        # ground_truth="ANOMALY→COMPOUND",
        # description="PE1 link-down and RR2 failure overlap; sessions lost then recovered.",
    # )
#     ft = _ft(3, "overlapping", "ld-rr-pe4-rr2")
#     writer.record_file(
#         filename=_fn("overlapping", "ld-rr-pe4-rr2"), section=3,
#         fault_type="link-down,rr-down", variant="overlapping",
#         affected_device="PE4,RR2",
#         fault_time=ft, recovery=True, recovery_time="~45s after fault",
#         ground_truth="ANOMALY→COMPOUND",
#         description="PE4 link-down and RR2 failure overlap; sessions lost then recovered.",
#     )

    # --- Link-Down + ESDF combos (2 files) ---
    # ft = _ft(3, "ld-esdf", "pe1-pe2")
    # writer.record_file(
        # filename=_fn("ld-esdf", "pe1-pe2"), section=3,
        # fault_type="link-down,esdf-toggle", variant="ld-esdf",
        # affected_device="PE1,PE2",
        # fault_time=ft, recovery=False,
        # ground_truth="ANOMALY→COMPOUND",
        # description="PE1 link-down combined with ESDF toggle on PE2; compound fault, no recovery.",
    # )
#     ft = _ft(3, "ld-esdf", "pe3-pe4")
#     writer.record_file(
#         filename=_fn("ld-esdf", "pe3-pe4"), section=3,
#         fault_type="link-down,esdf-toggle", variant="ld-esdf",
#         affected_device="PE3,PE4",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→COMPOUND",
#         description="PE3 link-down combined with ESDF toggle on PE4; compound fault, no recovery.",
#     )

    # --- Link-Down + RT-Misconfig combos (2 files) ---
    # ft = _ft(3, "ld-rt", "pe2-pe3")
    # writer.record_file(
        # filename=_fn("ld-rt", "pe2-pe3"), section=3,
        # fault_type="link-down,rt-misconfig", variant="ld-rt",
        # affected_device="PE2,PE3",
        # fault_time=ft, recovery=False,
        # ground_truth="ANOMALY→COMPOUND",
        # description="PE2 link-down concurrent with RT misconfiguration on PE3; compound silent+active fault.",
    # )
#     ft = _ft(3, "ld-rt", "pe4-pe5")
#     writer.record_file(
#         filename=_fn("ld-rt", "pe4-pe5"), section=3,
#         fault_type="link-down,rt-misconfig", variant="ld-rt",
#         affected_device="PE4,PE5",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→COMPOUND",
#         description="PE4 link-down concurrent with RT misconfiguration on PE5; compound silent+active fault.",
#     )

    # --- Planned Maintenance (7 files — pe1..pe5, rr1, rr2) ---
#     for pe in range(1, 6):
#         ft = _ft(3, "planned-maintenance", f"pe{pe}")
#         writer.record_file(
#             filename=_fn("planned-maintenance", f"pe{pe}"), section=3,
#             fault_type="planned-maintenance", variant=f"pe{pe}",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=True, recovery_time="~30s after fault",
#             ground_truth="NORMAL",
#             description=(
#                 f"PE{pe} performs a graceful shutdown: BGP NOTIFICATION (Cease/Admin Shutdown) "
#                 f"sent, session goes down cleanly, PE{pe} comes back up. Intentional operation."
#             ),
#         )
#     for rr in (1, 2):
#         ft = _ft(3, "planned-maintenance", f"rr{rr}")
#         writer.record_file(
#             filename=_fn("planned-maintenance", f"rr{rr}"), section=3,
#             fault_type="planned-maintenance", variant=f"rr{rr}",
#             affected_device=f"RR{rr}",
#             fault_time=ft, recovery=True, recovery_time="~30s after fault",
#             ground_truth="NORMAL",
#             description=(
#                 f"RR{rr} graceful shutdown: CEASE notification sent to all PEs, "
#                 f"then RR{rr} restarts and sessions recover. Intentional maintenance."
#             ),
#         )

    # --- Node Removal (2 files) ---
#     ft = _ft(3, "node-removal", "pe1")
#     writer.record_file(
#         filename=_fn("node-removal", "pe1"), section=3,
#         fault_type="node-removal", variant="pe1",
#         affected_device="PE1",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY",
#         description="PE1 permanently decommissioned mid-capture; TCP RST, session disappears from fabric.",
#     )
#     ft = _ft(3, "node-removal", "pe4")
#     writer.record_file(
#         filename=_fn("node-removal", "pe4"), section=3,
#         fault_type="node-removal", variant="pe4",
#         affected_device="PE4",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY",
#         description="PE4 permanently decommissioned mid-capture; TCP RST, session disappears from fabric.",
#     )

    # --- Unseen Topology (1 file) ---
#     ft = _ft(3, "unseen-topology", "pe6-joins")
#     writer.record_file(
#         filename=_fn("unseen-topology", "pe6-joins"), section=3,
#         fault_type="unseen-topology", variant="pe6-joins",
#         affected_device="PE6",
#         fault_time=ft, recovery=True,
#         ground_truth="ANOMALY",
#         description="New PE6 joins the fabric mid-capture with unknown device ID; tests generalisation to unseen topology.",
#     )

    # --- Cascade: RR-Down → ESDF (2 files) ---
#     ft = _ft(3, "cascade", "rr-down-esdf-rr1")
#     writer.record_file(
#         filename=_fn("cascade", "rr-down-esdf-rr1"), section=3,
#         fault_type="rr-down,esdf-toggle", variant="cascade",
#         affected_device="RR1",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→CASCADE",
#         description="RR1 failure cascades: loss of RR1 triggers ES/DF re-election on connected PEs.",
#     )
#     ft = _ft(3, "cascade", "rr-down-esdf-rr2")
#     writer.record_file(
#         filename=_fn("cascade", "rr-down-esdf-rr2"), section=3,
#         fault_type="rr-down,esdf-toggle", variant="cascade",
#         affected_device="RR2",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→CASCADE",
#         description="RR2 failure cascades: loss of RR2 triggers ES/DF re-election on connected PEs.",
#     )

#     # --- Cascade: Link-Down → RT-Misconfig (2 files) ---
#     ft = _ft(3, "cascade", "link-down-rtmisconfig-pe1")
#     writer.record_file(
#         filename=_fn("cascade", "link-down-rtmisconfig-pe1"), section=3,
#         fault_type="link-down,rt-misconfig", variant="cascade",
#         affected_device="PE1",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→CASCADE",
#         description="PE1 link-down followed by a cascading RT misconfiguration on the same device.",
#     )
#     ft = _ft(3, "cascade", "link-down-rtmisconfig-pe3")
#     writer.record_file(
#         filename=_fn("cascade", "link-down-rtmisconfig-pe3"), section=3,
#         fault_type="link-down,rt-misconfig", variant="cascade",
#         affected_device="PE3",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→CASCADE",
#         description="PE3 link-down followed by a cascading RT misconfiguration on the same device.",
#     )

    # --- Intermittent Link Flap (3 files) ---
#     for pe in (1, 2, 4):
#         ft = _ft(3, "intermittent", f"link-flap-pe{pe}")
#         writer.record_file(
#             filename=_fn("intermittent", f"link-flap-pe{pe}"), section=3,
#             fault_type="link-down", variant="intermittent",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=True,
#             ground_truth="ANOMALY→INTERMITTENT",
#             description=(
#                 f"PE{pe} link flaps repeatedly: multiple drop+reconnect cycles. "
#                 f"TCP RST each time → elevated OPEN count, no persistent failure."
#             ),
#         )

    # --- Intermittent ESDF Toggle (2 files) ---
#     for pe in (3, 5):
#         ft = _ft(3, "intermittent", f"esdf-pe{pe}")
#         writer.record_file(
#             filename=_fn("intermittent", f"esdf-pe{pe}"), section=3,
#             fault_type="esdf-toggle", variant="intermittent",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=True,
#             ground_truth="ANOMALY→INTERMITTENT",
#             description=(
#                 f"PE{pe} ESDF state flaps intermittently: EAD/ES routes toggled multiple times. "
#                 f"No session disruption, but high UPDATE churn."
#             ),
#         )

    # --- Session Flap (3 files) ---
#     ft = _ft(3, "session-flap", "pe1")
#     writer.record_file(
#         filename=_fn("session-flap", "pe1"), section=3,
#         fault_type="link-down", variant="session-flap",
#         affected_device="PE1",
#         fault_time=ft, recovery=True,
#         ground_truth="ANOMALY→INTERMITTENT",
#         description="PE1 BGP session flaps multiple times (OPEN=22, NOTIF=5): repeated teardown and re-establishment.",
#     )
#     ft = _ft(3, "session-flap", "pe2")
#     writer.record_file(
#         filename=_fn("session-flap", "pe2"), section=3,
#         fault_type="link-down", variant="session-flap",
#         affected_device="PE2",
#         fault_time=ft, recovery=True,
#         ground_truth="ANOMALY→INTERMITTENT",
#         description="PE2 BGP session flaps multiple times (OPEN=22, NOTIF=5): repeated teardown and re-establishment.",
#     )
#     ft = _ft(3, "session-flap", "rr1")
#     writer.record_file(
#         filename=_fn("session-flap", "rr1"), section=3,
#         fault_type="link-down", variant="session-flap",
#         affected_device="RR1",
#         fault_time=ft, recovery=True,
#         ground_truth="ANOMALY→INTERMITTENT",
#         description="RR1 BGP session flaps multiple times (OPEN=22, NOTIF=5): RR perspective of repeated flap.",
#     )

    # --- Slow Degradation (3 files) ---
#     for pe in (1, 2, 4):
#         ft = _ft(3, "slow-degradation", f"pe{pe}")
#         writer.record_file(
#             filename=_fn("slow-degradation", f"pe{pe}"), section=3,
#             fault_type="link-down", variant="slow-degradation",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=False,
#             ground_truth="ANOMALY→DEGRADATION",
#             description=(
#                 f"PE{pe} link quality slowly degrades: keepalive spacing increases until hold "
#                 f"timer expires (NOTIF=1). Short capture — progressive failure pattern."
#             ),
#         )

    # --- Mid-Session Link Down (3 files) ---
#     for pe in (1, 2, 3):
#         ft = _ft(3, "mid-session", f"link-down-pe{pe}")
#         writer.record_file(
#             filename=_fn("mid-session", f"link-down-pe{pe}"), section=3,
#             fault_type="link-down", variant="mid-session",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=False,
#             ground_truth="ANOMALY",
#             description=(
#                 f"PE{pe} link drops mid-capture after full route convergence (OPEN=2 only): "
#                 f"session silently disappears. Tests detection without session re-establishment context."
#             ),
#         )

    # --- AS Misconfiguration (3 files) ---
#     for pe in (1, 3, 5):
#         ft = _ft(3, "as-misconfig", f"pe{pe}")
#         writer.record_file(
#             filename=_fn("as-misconfig", f"pe{pe}"), section=3,
#             fault_type="as-misconfig", variant=f"pe{pe}",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=False,
#             ground_truth="ANOMALY",
#             description=(
#                 f"PE{pe} OPEN carries wrong AS number; peer rejects with NOTIFICATION "
#                 f"(code 2 Bad Peer AS). Session rejected on every re-attempt."
#             ),
#         )

    # --- Hold Timer Mismatch (2 files) ---
#     for pe in (2, 4):
#         ft = _ft(3, "hold-timer-mismatch", f"pe{pe}")
#         writer.record_file(
#             filename=_fn("hold-timer-mismatch", f"pe{pe}"), section=3,
#             fault_type="hold-timer-mismatch", variant=f"pe{pe}",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=False,
#             ground_truth="ANOMALY",
#             description=(
#                 f"PE{pe} OPEN advertises an incompatible hold time; negotiation fails → "
#                 f"BGP NOTIFICATION. Novel fault type not present in Section 2 training data."
#             ),
#         )

    # --- Max Prefix Limit (2 files) ---
#     for pe in (1, 4):
#         ft = _ft(3, "max-prefix", f"pe{pe}")
#         writer.record_file(
#             filename=_fn("max-prefix", f"pe{pe}"), section=3,
#             fault_type="max-prefix", variant=f"pe{pe}",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=False,
#             ground_truth="ANOMALY",
#             description=(
#                 f"PE{pe} exceeds configured max-prefix limit; RR sends NOTIFICATION (Cease/Max Prefixes). "
#                 f"High UPDATE count before the limit trip."
#             ),
#         )

    # --- Admin Reset (2 files) ---
#     for pe in (2, 3):
#         ft = _ft(3, "admin-reset", f"pe{pe}")
#         writer.record_file(
#             filename=_fn("admin-reset", f"pe{pe}"), section=3,
#             fault_type="admin-reset", variant=f"pe{pe}",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=True, recovery_time="~30s after fault",
#             ground_truth="ANOMALY",
#             description=(
#                 f"PE{pe} session administratively reset mid-capture: NOTIFICATION (Cease/Admin Reset), "
#                 f"session torn down, then recovers. Looks like planned but without graceful GR."
#             ),
#         )

    # --- Peer Deconfiguration (2 files) ---
#     for pe in (1, 5):
#         ft = _ft(3, "peer-deconfig", f"pe{pe}")
#         writer.record_file(
#             filename=_fn("peer-deconfig", f"pe{pe}"), section=3,
#             fault_type="peer-deconfig", variant=f"pe{pe}",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=False,
#             ground_truth="ANOMALY",
#             description=(
#                 f"PE{pe} is deconfigured as a BGP peer: NOTIFICATION (Cease/Peer Deconfigured) "
#                 f"sent, session permanently removed."
#             ),
#         )

    # --- Invalid Next-Hop (2 files) ---
#     for pe in (1, 3):
#         ft = _ft(3, "invalid-nexthop", f"pe{pe}")
#         writer.record_file(
#             filename=_fn("invalid-nexthop", f"pe{pe}"), section=3,
#             fault_type="invalid-nexthop", variant=f"pe{pe}",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=False,
#             ground_truth="ANOMALY",
#             description=(
#                 f"PE{pe} advertises EVPN routes with an unreachable next-hop IPv6 address. "
#                 f"Routes accepted at BGP level but silently unusable. NOTIFICATION=1."
#             ),
#         )

    # --- Duplicate MAC (2 files) ---
#     ft = _ft(3, "dup-mac", "pe1-pe3")
#     writer.record_file(
#         filename=_fn("dup-mac", "pe1-pe3"), section=3,
#         fault_type="dup-mac", variant="pe1-pe3",
#         affected_device="PE1,PE3",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY",
#         description="PE1 and PE3 both advertise the same MAC in EVPN Type-2 routes; silent data-plane conflict.",
#     )
#     ft = _ft(3, "dup-mac", "pe2-pe4")
#     writer.record_file(
#         filename=_fn("dup-mac", "pe2-pe4"), section=3,
#         fault_type="dup-mac", variant="pe2-pe4",
#         affected_device="PE2,PE4",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY",
#         description="PE2 and PE4 both advertise the same MAC in EVPN Type-2 routes; silent data-plane conflict.",
#     )

    # --- VNI Mismatch (2 files) ---
#     for pe in (2, 4):
#         ft = _ft(3, "vni-mismatch", f"pe{pe}")
#         writer.record_file(
#             filename=_fn("vni-mismatch", f"pe{pe}"), section=3,
#             fault_type="vni-mismatch", variant=f"pe{pe}",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=False,
#             ground_truth="ANOMALY",
#             description=(
#                 f"PE{pe} advertises EVPN routes with an incorrect VNI in NLRI. "
#                 f"Routes structurally valid but semantically wrong; silent fabric isolation."
#             ),
#         )

    # --- FSM Error (2 files) ---
#     for pe in (1, 3):
#         ft = _ft(3, "fsm-error", f"pe{pe}")
#         writer.record_file(
#             filename=_fn("fsm-error", f"pe{pe}"), section=3,
#             fault_type="fsm-error", variant=f"pe{pe}",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=False,
#             ground_truth="ANOMALY",
#             description=(
#                 f"PE{pe} sends an unexpected BGP message in the wrong FSM state → "
#                 f"NOTIFICATION (FSM Error). Session terminated."
#             ),
#         )

    # --- Malformed AS-Path (2 files) ---
#     for pe in (2, 5):
#         ft = _ft(3, "malformed-aspath", f"pe{pe}")
#         writer.record_file(
#             filename=_fn("malformed-aspath", f"pe{pe}"), section=3,
#             fault_type="malformed-aspath", variant=f"pe{pe}",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=False,
#             ground_truth="ANOMALY",
#             description=(
#                 f"PE{pe} sends UPDATE with malformed AS_PATH attribute → "
#                 f"NOTIFICATION (Update Error/Malformed AS_PATH). Session terminated."
#             ),
#         )

    # --- Out of Resources (2 files) ---
#     for rr in (1, 2):
#         ft = _ft(3, "out-of-resources", f"rr{rr}")
#         writer.record_file(
#             filename=_fn("out-of-resources", f"rr{rr}"), section=3,
#             fault_type="out-of-resources", variant=f"rr{rr}",
#             affected_device=f"RR{rr}",
#             fault_time=ft, recovery=False,
#             ground_truth="ANOMALY",
#             description=(
#                 f"RR{rr} exhausts resources: sends multiple NOTIFICATION (Cease/Out of Resources) "
#                 f"messages terminating all sessions (NOTIF=6). Very short capture ~900 frames."
#             ),
#         )

    # --- RR-Down + ESDF pairwise (2 files) ---
#     ft = _ft(3, "rr-esdf", "rr1-pe3")
#     writer.record_file(
#         filename=_fn("rr-esdf", "rr1-pe3"), section=3,
#         fault_type="rr-down,esdf-toggle", variant="rr-esdf",
#         affected_device="RR1,PE3",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→COMPOUND",
#         description="RR1 failure combined with ESDF toggle on PE3; compound fault not in Section 2.",
#     )
#     ft = _ft(3, "rr-esdf", "rr2-pe5")
#     writer.record_file(
#         filename=_fn("rr-esdf", "rr2-pe5"), section=3,
#         fault_type="rr-down,esdf-toggle", variant="rr-esdf",
#         affected_device="RR2,PE5",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→COMPOUND",
#         description="RR2 failure combined with ESDF toggle on PE5; compound fault not in Section 2.",
#     )

    # --- RR-Down + RT-Misconfig pairwise (2 files) ---
    # ft = _ft(3, "rr-rt", "rr1-pe2")
    # writer.record_file(
        # filename=_fn("rr-rt", "rr1-pe2"), section=3,
        # fault_type="rr-down,rt-misconfig", variant="rr-rt",
        # affected_device="RR1,PE2",
        # fault_time=ft, recovery=False,
        # ground_truth="ANOMALY→COMPOUND",
        # description="RR1 down concurrent with RT misconfiguration on PE2; compound fault not in Section 2.",
    # )
#     ft = _ft(3, "rr-rt", "rr2-pe4")
#     writer.record_file(
#         filename=_fn("rr-rt", "rr2-pe4"), section=3,
#         fault_type="rr-down,rt-misconfig", variant="rr-rt",
#         affected_device="RR2,PE4",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→COMPOUND",
#         description="RR2 down concurrent with RT misconfiguration on PE4; compound fault not in Section 2.",
#     )

    # --- ESDF + RT-Misconfig pairwise (2 files) ---
    # ft = _ft(3, "esdf-rt", "pe1-pe2")
    # writer.record_file(
        # filename=_fn("esdf-rt", "pe1-pe2"), section=3,
        # fault_type="esdf-toggle,rt-misconfig", variant="esdf-rt",
        # affected_device="PE1,PE2",
        # fault_time=ft, recovery=False,
        # ground_truth="ANOMALY→COMPOUND",
        # description="ESDF toggle on PE1 combined with RT misconfiguration on PE2; novel pairwise combination.",
    # )
#     ft = _ft(3, "esdf-rt", "pe3-pe4")
#     writer.record_file(
#         filename=_fn("esdf-rt", "pe3-pe4"), section=3,
#         fault_type="esdf-toggle,rt-misconfig", variant="esdf-rt",
#         affected_device="PE3,PE4",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→COMPOUND",
#         description="ESDF toggle on PE3 combined with RT misconfiguration on PE4; novel pairwise combination.",
#     )

    # --- Triple combinations (4 files) ---
    # ft = _ft(3, "triple", "ld-rr-esdf")
    # writer.record_file(
        # filename=_fn("triple", "ld-rr-esdf"), section=3,
        # fault_type="link-down,rr-down,esdf-toggle", variant="triple",
        # affected_device="PE1,RR2,PE3",
        # fault_time=ft, recovery=False,
        # ground_truth="ANOMALY→CASCADE",
        # description="Three simultaneous fault types: link-down + RR-down + ESDF toggle. Tests multi-fault isolation.",
    # )
#     ft = _ft(3, "triple", "ld-rr-rt")
#     writer.record_file(
#         filename=_fn("triple", "ld-rr-rt"), section=3,
#         fault_type="link-down,rr-down,rt-misconfig", variant="triple",
#         affected_device="PE2,RR1,PE4",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→CASCADE",
#         description="Three simultaneous fault types: link-down + RR-down + RT-misconfig.",
#     )
#     ft = _ft(3, "triple", "ld-esdf-rt")
#     writer.record_file(
#         filename=_fn("triple", "ld-esdf-rt"), section=3,
#         fault_type="link-down,esdf-toggle,rt-misconfig", variant="triple",
#         affected_device="PE1,PE3,PE5",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→CASCADE",
#         description="Three simultaneous fault types: link-down + ESDF toggle + RT-misconfig.",
#     )
#     ft = _ft(3, "triple", "rr-esdf-rt")
#     writer.record_file(
#         filename=_fn("triple", "rr-esdf-rt"), section=3,
#         fault_type="rr-down,esdf-toggle,rt-misconfig", variant="triple",
#         affected_device="RR1,PE2,PE4",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→CASCADE",
#         description="Three simultaneous fault types: RR-down + ESDF toggle + RT-misconfig.",
#     )

    # --- Cross: existing + new fault type combos (5 files) ---
#     ft = _ft(3, "cross", "ld-max-prefix")
#     writer.record_file(
#         filename=_fn("cross", "ld-max-prefix"), section=3,
#         fault_type="link-down,max-prefix", variant="cross",
#         affected_device="PE1,PE4",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→COMPOUND",
#         description="Link-down (known) combined with max-prefix limit breach (novel); compound with new fault type.",
#     )
#     ft = _ft(3, "cross", "rr-as-misconfig")
#     writer.record_file(
#         filename=_fn("cross", "rr-as-misconfig"), section=3,
#         fault_type="rr-down,as-misconfig", variant="cross",
#         affected_device="RR1,PE3",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→COMPOUND",
#         description="RR-down (known) combined with AS misconfiguration (novel); compound with new fault type.",
#     )
#     ft = _ft(3, "cross", "esdf-dup-mac")
#     writer.record_file(
#         filename=_fn("cross", "esdf-dup-mac"), section=3,
#         fault_type="esdf-toggle,dup-mac", variant="cross",
#         affected_device="PE1,PE3",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→COMPOUND",
#         description="ESDF toggle (known) combined with duplicate MAC advertisement (novel).",
#     )
#     ft = _ft(3, "cross", "rt-invalid-nexthop")
#     writer.record_file(
#         filename=_fn("cross", "rt-invalid-nexthop"), section=3,
#         fault_type="rt-misconfig,invalid-nexthop", variant="cross",
#         affected_device="PE2,PE4",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→COMPOUND",
#         description="RT misconfiguration (known) combined with invalid next-hop (novel); both silent faults.",
#     )
#     ft = _ft(3, "cross", "ld-rr-max-prefix")
#     writer.record_file(
#         filename=_fn("cross", "ld-rr-max-prefix"), section=3,
#         fault_type="link-down,rr-down,max-prefix", variant="cross",
#         affected_device="PE2,RR2,PE1",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→CASCADE",
#         description="Link-down + RR-down cascade (known combo) plus max-prefix breach (novel); triple cross-combination.",
#     )

    # =========================================================================
    # SECTION 4 — Additional temporal / cascade scenarios (21 files)
    # =========================================================================

    # --- Cascade: RR-Down → ESDF (2 files) ---
    # ft = _ft(4, "cascade", "rr-down-esdf-rr1")
    # writer.record_file(
        # filename=_fn("cascade", "rr-down-esdf-rr1"), section=4,
        # fault_type="rr-down,esdf-toggle", variant="cascade",
        # affected_device="RR1",
        # fault_time=ft, recovery=False,
        # ground_truth="ANOMALY→CASCADE",
        # description="RR1 failure triggers cascading ES/DF re-election on connected PEs.",
    # )
    # ft = _ft(4, "cascade", "rr-down-esdf-rr2")
    # writer.record_file(
        # filename=_fn("cascade", "rr-down-esdf-rr2"), section=4,
        # fault_type="rr-down,esdf-toggle", variant="cascade",
        # affected_device="RR2",
        # fault_time=ft, recovery=False,
        # ground_truth="ANOMALY→CASCADE",
        # description="RR2 failure triggers cascading ES/DF re-election on connected PEs.",
    # )

    # --- Cascade: Link-Down → RT-Misconfig (2 files) ---
    # ft = _ft(4, "cascade", "link-down-rtmisconfig-pe1")
    # writer.record_file(
        # filename=_fn("cascade", "link-down-rtmisconfig-pe1"), section=4,
        # fault_type="link-down,rt-misconfig", variant="cascade",
        # affected_device="PE1",
        # fault_time=ft, recovery=False,
        # ground_truth="ANOMALY→CASCADE",
        # description="PE1 link-down triggers a cascading RT misconfiguration on the same device.",
    # )
#     ft = _ft(4, "cascade", "link-down-rtmisconfig-pe3")
#     writer.record_file(
#         filename=_fn("cascade", "link-down-rtmisconfig-pe3"), section=4,
#         fault_type="link-down,rt-misconfig", variant="cascade",
#         affected_device="PE3",
#         fault_time=ft, recovery=False,
#         ground_truth="ANOMALY→CASCADE",
#         description="PE3 link-down triggers a cascading RT misconfiguration on the same device.",
#     )

    # --- Intermittent Link Flap (3 files) ---
    # for pe in (1, 2):
        # ft = _ft(4, "intermittent", f"link-flap-pe{pe}")
        # writer.record_file(
            # filename=_fn("intermittent", f"link-flap-pe{pe}"), section=4,
            # fault_type="link-down", variant="intermittent",
            # affected_device=f"PE{pe}",
            # fault_time=ft, recovery=True,
            # ground_truth="ANOMALY→INTERMITTENT",
            # description=(
                # f"PE{pe} link flaps repeatedly across the full capture window. "
                # f"Elevated OPEN count (24); tests long-horizon intermittent fault detection."
            # ),
        # )

    # --- Intermittent ESDF (2 files) ---
#     for pe in (3, 5):
#         ft = _ft(4, "intermittent", f"esdf-pe{pe}")
#         writer.record_file(
#             filename=_fn("intermittent", f"esdf-pe{pe}"), section=4,
#             fault_type="esdf-toggle", variant="intermittent",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=True,
#             ground_truth="ANOMALY→INTERMITTENT",
#             description=(
#                 f"PE{pe} ESDF state flaps across the full capture; repeated EAD/ES churn "
#                 f"without any BGP session disruption."
#             ),
#         )

    # --- Slow Degradation (3 files) ---
#     for pe in (1, 2, 4):
#         ft = _ft(4, "slow-degradation", f"pe{pe}")
#         writer.record_file(
#             filename=_fn("slow-degradation", f"pe{pe}"), section=4,
#             fault_type="link-down", variant="slow-degradation",
#             affected_device=f"PE{pe}",
#             fault_time=ft, recovery=False,
#             ground_truth="ANOMALY→DEGRADATION",
#             description=(
#                 f"PE{pe} link quality degrades progressively: keepalive spacing widens until "
#                 f"hold timer fires (NOTIF=1). Short capture representing the degradation period."
#             ),
#         )

    # --- Session Flap (3 files) ---
#     ft = _ft(4, "session-flap", "pe1")
#     writer.record_file(
#         filename=_fn("session-flap", "pe1"), section=4,
#         fault_type="link-down", variant="session-flap",
#         affected_device="PE1",
#         fault_time=ft, recovery=True,
#         ground_truth="ANOMALY→INTERMITTENT",
#         description="PE1 BGP session flaps multiple times (OPEN=22, NOTIF=5) across the full capture.",
#     )
#     ft = _ft(4, "session-flap", "pe2")
#     writer.record_file(
#         filename=_fn("session-flap", "pe2"), section=4,
#         fault_type="link-down", variant="session-flap",
#         affected_device="PE2",
#         fault_time=ft, recovery=True,
#         ground_truth="ANOMALY→INTERMITTENT",
#         description="PE2 BGP session flaps multiple times (OPEN=22, NOTIF=5) across the full capture.",
#     )
#     ft = _ft(4, "session-flap", "rr1")
#     writer.record_file(
#         filename=_fn("session-flap", "rr1"), section=4,
#         fault_type="link-down", variant="session-flap",
#         affected_device="RR1",
#         fault_time=ft, recovery=True,
#         ground_truth="ANOMALY→INTERMITTENT",
#         description="RR1 perspective of repeated session flap (OPEN=22, NOTIF=5); tests RR-side detection.",
#     )

    # --- RT-Misconfig with Recovery (3 files) ---
    # for pe in (1, 2, 4):
        # ft = _ft(2, "rt-misconfig", f"recovery-pe{pe}")
        # writer.record_file(
            # filename=_fn("rt-misconfig-recovery", f"pe{pe}"), section=4,
            # fault_type="rt-misconfig", variant="recovery",
            # affected_device=f"PE{pe}",
            # fault_time=ft, recovery=True, recovery_time="~60s after fault",
            # ground_truth="ANOMALY→RECOVERY",
            # description=(
                # f"PE{pe} starts with wrong RT export; misconfiguration corrected mid-capture "
                # f"and routes re-advertised with the correct RT. ~20K frames."
            # ),
        # )

    # --- Mid-Session Link Down (3 files) ---
    # for pe in (1, 2, 3):
        # ft = _ft(4, "mid-session", f"link-down-pe{pe}")
        # writer.record_file(
            # filename=_fn("mid-session", f"link-down-pe{pe}"), section=4,
            # fault_type="link-down", variant="mid-session",
            # affected_device=f"PE{pe}",
            # fault_time=ft, recovery=False,
            # ground_truth="ANOMALY",
            # description=(
                # f"PE{pe} link drops after full convergence (OPEN=2 only): session silently "
                # f"disappears mid-capture. Tests detection without prior session context."
            # ),
        # )

    # =========================================================================
    # Additional classes, sourced from scripts/generate_json.py's CATALOGUE
    # =========================================================================
    _session_added_keys = [
        ("section2_labelled", "esdf_toggle_single_midchurn_pe1"),
        ("section2_labelled", "esdf_toggle_single_midchurn_pe2"),
        # ("section3_mixed", "ld_triggers_esdf_pe1"),
        # ("section3_mixed", "ld_triggers_esdf_pe2"),
        # ("section3_mixed", "ld_esdf_overlap_pe1_pe2"),
        # ("section3_mixed", "ld_esdf_overlap_pe3_pe2"),
        # ("section3_mixed", "ld_rt_overlap_pe2_pe3"),
        # ("section3_mixed", "ld_rt_overlap_pe3_pe1"),
        # ("section3_mixed", "rr_then_ld_rr2_pe1"),
        # ("section3_mixed", "rr_then_ld_rr2_pe3"),
    ]
    try:
        import importlib.util as _ilu
        _gj_path = Path(__file__).parent.parent.parent / "scripts" / "generate_json.py"
        _spec = _ilu.spec_from_file_location("generate_json", _gj_path)
        _gj = _ilu.module_from_spec(_spec)
        _spec.loader.exec_module(_gj)
        for _sec_dir, _slug in _session_added_keys:
            _entry = _gj.CATALOGUE.get((_sec_dir, _slug))
            if _entry is None:
                continue
            writer.record_file(
                filename=f"{_slug}.pcap", section=_entry["section"],
                fault_type=_entry["fault_type"], variant=_slug,
                affected_device=_entry["affected_device"],
                fault_time=_entry["fault_inject_time_seconds"],
                recovery=_entry["recovery"],
                recovery_time=_entry["recovery_time_seconds"],
                ground_truth=_entry["ground_truth_label"],
                description=_entry["description"],
            )
    except Exception as e:
        import warnings
        warnings.warn(
            f"Failed to load session-added metadata rows from generate_json.py's "
            f"CATALOGUE ({e}). dataset_metadata.xlsx will be missing these rows.",
            stacklevel=2,
        )

    return writer
